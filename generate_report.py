#! /usr/bin/python

import os
import sys
import time
import importlib
import opendr
import numpy as np
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import (_get_column_letter)
from src.measure import Measure, getMeasures

if __name__ == '__main__':
	
	if len(sys.argv) < 3:
		print 'Please provide as an argument the directory where the OBJ-files are located and the shape to be used'
		sys.exit(1)
	
	objModels = []
	for file in os.listdir(sys.argv[1]):
		if file.endswith('.obj'):
			objModels.append(file)
	
	wb = Workbook()
	ws = wb.active

	ws.column_dimensions['A'].width = 25

	module = importlib.import_module('src.shapes')
	class_ = getattr(module, sys.argv[2])

	#measures = getMeasures(Sphere)
	measures = getMeasures(class_)
	rule = ColorScaleRule(start_type='percentile', start_value=10, start_color='00FF00', mid_type='percentile', mid_value=50, mid_color='FFFF00', end_type='percentile', end_value=90, end_color='FF0000')

	j = 1
	for measure in measures:
		c = ws.cell(row = 1, column = j)
		if type(measure.name) is list:
			for submeasure in measure.name:
				if measure.color:
					ws.conditional_formatting.add(_get_column_letter(j) + '2:' + _get_column_letter(j) + str(len(objModels)+1), rule)
				c = ws.cell(row = 1, column = j)
				c.value = submeasure
				j += 1
		else:
			if measure.color:
				ws.conditional_formatting.add(_get_column_letter(j) + '2:' + _get_column_letter(j) + str(len(objModels)+1), rule)
			c.value = measure.name
			j += 1

	for i,fileName in enumerate(objModels):
		print 'Processing file ' + str(i+1) + ' of ' + str(len(objModels)) + '..'
		measurements = dict()
		print fileName
		
		filePath = sys.argv[1] + '/' + fileName

		#shapeObj = Sphere(filePath, 39.5)
		shapeObj = class_(filePath, *sys.argv[3:])

		j = 1
		for index, measurement in enumerate(measures):
			res = measurement.execute(shapeObj)
			if type(res) is list:
				for r in res:
					c = ws.cell(row = i + 2, column = j)
					c.value = r
					j += 1
			else:
				c = ws.cell(row = i + 2, column = j)
				c.value = res
				j += 1

	wb.save('report.xlsx')
