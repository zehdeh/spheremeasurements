from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import (_get_column_letter)
import csv

def writeReport(filePath, measures, results, write_csv):
	wb = Workbook()
	ws = wb.active

	rule = ColorScaleRule(start_type='percentile', start_value=10, start_color='00FF00', mid_type='percentile', mid_value=50, mid_color='FFFF00', end_type='percentile', end_value=90, end_color='FF0000')

	for i,measure in enumerate(measures):
		c = ws.cell(row = 1, column = i+2)
		c.value = measure.name
		if measure.color and not write_csv:
			ws.conditional_formatting.add(_get_column_letter(i+2) + '2:' + _get_column_letter(i+2) + str(len(results)+1), rule)
	
	for i, result in enumerate(results):
		for j, val in enumerate(result):
			c = ws.cell(row = 2+i, column = j+1)
			try:
				c.value = val
				if j > 0 and measures[j-1].formatInteger:
					c.number_format = '0'
			except ValueError:
				c.value = float(val)

	if write_csv:
		sh = wb.active
		with open(filePath, 'wb') as f:
			c = csv.writer(f)
			for i,r in enumerate(sh.rows):
				row = [cell.value for cell in r]
				if i > 0:
					for j, val in enumerate(row):
						if j > 0 and measures[j-1].formatInteger:
							row[j] = int(val)
				c.writerow(row)
	else:
		wb.save(filePath)

